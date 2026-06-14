"""Secondary-school selectivity from CERMAT unified entrance exam (JPZ).

Source: data.cermat.cz aggregated school-level results. Per school-field
we get the average percentile of applicants in Czech+Math (ČJ+MA). A
higher percentile means the school attracts higher-performing pupils — a
common proxy for school quality/selectivity. Open data.
"""
import io

import requests
from openpyxl import load_workbook

RESULTS_URL = (
    "https://data.cermat.cz/files/files/JPZ/agregovana_data_skoly/"
    "PZ2026_kolo1_skolobory_vysledky.xlsx"
)

# column indices in the results sheet
_IZO = 4
_PERCENTILE = 57   # ČJ+MA - PERCENTIL - PRŮMĚR
_TOOK = 45         # ČJ+MA - KONALI (number of pupils, used as weight)

_percentiles: dict[str, float] | None = None


def _norm_izo(raw) -> str:
    """'izo_000013757' → '13757' (registry IZO has no leading zeros)."""
    return str(raw).replace("izo_", "").lstrip("0")


def school_percentiles() -> dict[str, float]:
    """Return {normalized_izo: weighted-mean ČJ+MA percentile (0-100)}."""
    global _percentiles
    if _percentiles is not None:
        return _percentiles

    r = requests.get(RESULTS_URL, timeout=300)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    sums: dict[str, float] = {}
    weights: dict[str, float] = {}
    for row in rows:
        perc = row[_PERCENTILE]
        if perc is None:
            continue
        try:
            perc = float(perc)
            w = float(row[_TOOK] or 0) or 1.0
        except (ValueError, TypeError):
            continue
        izo = _norm_izo(row[_IZO])
        sums[izo] = sums.get(izo, 0.0) + perc * w
        weights[izo] = weights.get(izo, 0.0) + w

    _percentiles = {izo: sums[izo] / weights[izo] for izo in sums}
    return _percentiles
